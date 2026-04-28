"""
HỆ THỐNG TỰ ĐỘNG HÓA BIÊN BẢN DƯỢC – Bệnh viện Đà Nẵng
Hỗ trợ: Biên Bản Kiểm Nhập (BBKN) & Báo Cáo Xuất Nhập Tồn (XNT)
"""

import io, math, copy, datetime, warnings
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Biên Bản Dược – BV Đà Nẵng",
    page_icon="🏥",
    layout="centered",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Be+Vietnam+Pro:wght@300;400;600;700&display=swap');
html,body,[class*="css"]{font-family:'Be Vietnam Pro',sans-serif;}
.hero{background:linear-gradient(135deg,#1a3a5c 0%,#2563a8 60%,#1e7fcb 100%);
  border-radius:16px;padding:32px 36px 24px;margin-bottom:24px;color:white;
  box-shadow:0 8px 32px rgba(37,99,168,.25);}
.hero h1{font-size:1.55rem;font-weight:700;margin:0 0 6px;line-height:1.3;}
.hero .sub{font-size:.88rem;font-weight:300;opacity:.85;margin:0;}
.hero .badge{display:inline-block;background:rgba(255,255,255,.18);border-radius:20px;
  padding:3px 12px;font-size:.75rem;font-weight:600;letter-spacing:1px;
  margin-bottom:12px;text-transform:uppercase;}
.tab-desc{background:#eff6ff;border-left:4px solid #2563a8;border-radius:0 10px 10px 0;
  padding:12px 16px;margin:12px 0 18px;font-size:.86rem;color:#1e3a5f;line-height:1.6;}
.stat-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin:16px 0;}
.stat-card{background:white;border:1px solid #e2e8f0;border-radius:12px;
  padding:16px 12px;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,.06);}
.stat-card .num{font-size:1.7rem;font-weight:700;color:#1a3a5c;line-height:1;}
.stat-card .lbl{font-size:.75rem;color:#64748b;margin-top:4px;}
.ok-box{background:#f0fdf4;border:1.5px solid #86efac;border-radius:12px;
  padding:18px 20px;margin:16px 0;text-align:center;}
.ok-box .icon{font-size:2rem;}.ok-box h3{color:#166534;margin:6px 0 4px;font-size:1rem;}
.ok-box p{color:#15803d;font-size:.84rem;margin:0;}
.note{background:#fff7ed;border-left:4px solid #f59e0b;border-radius:0 10px 10px 0;
  padding:10px 14px;font-size:.82rem;color:#92400e;margin-top:14px;line-height:1.6;}
.stButton>button{background:linear-gradient(135deg,#1a3a5c,#2563a8)!important;
  color:white!important;font-weight:600!important;font-size:.95rem!important;
  border:none!important;border-radius:10px!important;padding:13px 0!important;
  width:100%!important;box-shadow:0 4px 14px rgba(37,99,168,.3)!important;}
[data-testid="stDownloadButton"]>button{background:linear-gradient(135deg,#166534,#16a34a)!important;
  color:white!important;font-weight:700!important;font-size:1rem!important;
  border:none!important;border-radius:10px!important;padding:15px 0!important;
  width:100%!important;box-shadow:0 4px 14px rgba(22,163,74,.3)!important;}
[data-testid="stFileUploader"]{border:2px dashed #2563a8!important;
  border-radius:12px!important;background:#f0f6ff!important;}
hr{border:none;border-top:1px solid #e2e8f0;margin:20px 0;}
#MainMenu,footer{visibility:hidden;}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
#  SHARED HELPERS
# ══════════════════════════════════════════════════════════════════════════════
SKIP_KW = ['Tổng cộng','Hội đồng','Trưởng','Trang','Đã kiểm nhập',
           'Ông/bà','kiểm nhập những','Trang 1']
THIN = Side(style='thin')
MED  = Side(style='medium')
NO_FILL = PatternFill(fill_type=None)

def b_thin(): return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
def b_med():  return Border(left=THIN, right=THIN, top=MED,  bottom=MED)

def safe_set(cell, **kwargs):
    for k, v in kwargs.items():
        try: setattr(cell, k, v)
        except AttributeError: pass

def is_co_row(v0, row):
    return (isinstance(v0, str) and pd.isna(row[1]) and pd.isna(row[2])
            and not any(kw in str(v0) for kw in SKIP_KW))

def is_drug_row(v0, col2):
    try: int(str(v0).strip())
    except: return False
    return not pd.isna(col2) and isinstance(col2, str) and not col2.strip().isdigit()

def parse_companies(raw_df, qty_col):
    result, cur, rows, skipped = [], None, [], 0
    for _, row in raw_df.iterrows():
        v0 = row[0]
        if is_co_row(v0, row):
            if cur and rows: result.append((cur, rows))
            cur, rows = str(v0).strip(), []
        elif is_drug_row(v0, row[2]):
            try:    qty = float(row[qty_col]) if not pd.isna(row[qty_col]) else 0
            except: qty = 0
            if qty != 0: rows.append(row)
            else: skipped += 1
    if cur and rows: result.append((cur, rows))
    return result, {'companies': len(result),
                    'drugs': sum(len(d) for _, d in result),
                    'skipped': skipped}

def gs(ws, r, c):
    cl = ws.cell(row=r, column=c)
    return {k: copy.copy(getattr(cl, k))
            for k in ('font','border','alignment','fill','number_format')}

def ap(cell, s):
    for k, v in s.items():
        try: setattr(cell, k, copy.copy(v))
        except AttributeError: pass


# ══════════════════════════════════════════════════════════════════════════════
#  MODULE BBKN
# ══════════════════════════════════════════════════════════════════════════════
BBKN_W = {1:5.5,2:10,3:38,4:17,5:7.5,6:11,7:24,8:11,9:12,10:9.5,11:14,12:8}
BBKN_A = {1:('center','center'),2:('center','center'),3:('left','center'),
          4:('left','center'),5:('center','center'),6:('center','center'),
          7:('left','center'),8:('center','center'),9:('right','center'),
          10:('right','center'),11:('right','center'),12:('center','center')}
BBKN_WRAP = {3,4,7}
BBKN_NUM  = {9,10,11}

def bbkn_h(ws, r):
    ml = 1
    for c in BBKN_WRAP:
        v = ws.cell(row=r, column=c).value
        if not v or not isinstance(v, str): continue
        cw = max(BBKN_W.get(c,15)*1.1, 1)
        ml = max(ml, sum(max(1,math.ceil(len(ln)/cw)) for ln in v.split('\n')))
    return max(22, min(ml*15.6+4, 120))

def build_bbkn(tmpl_bytes, companies):
    wb = load_workbook(io.BytesIO(tmpl_bytes))
    ws = wb.active
    cs = {c: gs(ws,15,c) for c in range(1,13)}
    ds = {c: gs(ws,16,c) for c in range(1,13)}
    tks = gs(ws,213,11)

    fs = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == 'HỘI ĐỒNG KIỂM NHẬP': fs = cell.row; break
        if fs: break
    if not fs: fs = 215

    DS = 15
    need = sum(1+len(d) for _,d in companies)+1
    ins = (DS+need-1) - fs + 1
    if ins > 0: ws.insert_rows(fs, ins); fs += ins

    for m in [str(mr) for mr in ws.merged_cells.ranges if DS <= mr.min_row < fs]:
        ws.merged_cells.remove(m)
    for r in range(DS, fs):
        for c in range(1,13):
            try: ws.cell(row=r,column=c).value = None
            except: pass

    def wco(rn, name):
        cl = ws.cell(row=rn, column=1, value=name); ap(cl, cs[1])
        cl.font = Font(name='Times New Roman', bold=True, size=12); cl.fill = NO_FILL
        for c in range(2,13):
            cc = ws.cell(row=rn, column=c); ap(cc, cs[c]); cc.fill = NO_FILL
        ws.row_dimensions[rn].height = 20

    def wdr(rn, stt, dr):
        cols = [
            (1, stt,                                               'center', False, None),
            (2, ''if pd.isna(dr[1])else str(dr[1]).strip(),       'center', False, None),
            (3, ''if pd.isna(dr[2])else str(dr[2]).strip(),       'left',   True,  None),
            (4, ''if pd.isna(dr[3])else str(dr[3]).strip(),       'left',   True,  None),
            (5, ''if pd.isna(dr[4])else str(dr[4]).strip(),       'center', False, None),
            (6, ''if pd.isna(dr[5])else str(dr[5]).strip(),       'center', False, None),
            (7, ''if pd.isna(dr[6])else str(dr[6]).strip(),       'left',   True,  None),
            (8, dr[7] if isinstance(dr[7],datetime.datetime)
                else(''if pd.isna(dr[7])else dr[7]),              'center', False, 'DD/MM/YYYY'),
            (9, dr[8] if not pd.isna(dr[8]) else 0,              'right',  False, '#,##0'),
            (10,int(dr[9]) if not pd.isna(dr[9]) else 0,         'right',  False, '#,##0'),
        ]
        for col,val,ha,wrap,fmt in cols:
            cl = ws.cell(row=rn,column=col,value=val); ap(cl,ds[col])
            cl.font = Font(name='Times New Roman',size=12)
            cl.alignment = Alignment(horizontal=ha,vertical='center',wrap_text=wrap)
            if fmt and val!='': cl.number_format = fmt
        ck = ws.cell(row=rn,column=11,value=f'=I{rn}*J{rn}'); ap(ck,ds[11])
        ck.font=Font(name='Times New Roman',size=12)
        ck.alignment=Alignment(horizontal='right',vertical='center')
        ck.number_format='#,##0'
        cl12=ws.cell(row=rn,column=12,value=''); ap(cl12,ds[12])

    cr = DS; drn = []
    for name,drugs in companies:
        wco(cr,name); cr+=1
        for i,dr in enumerate(drugs,1): wdr(cr,i,dr); drn.append(cr); cr+=1

    tr = cr
    lbl=ws.cell(row=tr,column=1,value='Tổng cộng: ')
    lbl.font=Font(name='Times New Roman',bold=True,size=12)
    lbl.alignment=Alignment(horizontal='left',vertical='center'); lbl.border=b_med()
    for c in range(2,11):
        try: ws.cell(row=tr,column=c).border=b_med()
        except: pass
    ck=ws.cell(row=tr,column=11,value=f'=SUM({",".join(f"K{r}"for r in drn)})')
    ap(ck,tks); ck.font=Font(name='Times New Roman',bold=True,size=12)
    ck.alignment=Alignment(horizontal='right',vertical='center')
    ck.number_format='#,##0'; ck.border=b_med()
    ws.row_dimensions[tr].height=22

    ws.cell(row=13,column=3).value='Tên thuốc'
    for col in range(1,13):
        for r in (13,14):
            safe_set(ws.cell(row=r,column=col),fill=NO_FILL,
                     font=Font(name='Times New Roman',bold=True,size=12),
                     border=b_med(),
                     alignment=Alignment(horizontal='center',vertical='center',wrap_text=True))
    ws.row_dimensions[13].height=42; ws.row_dimensions[14].height=18

    for r in range(DS,tr+1):
        av=ws.cell(row=r,column=1).value; cv=ws.cell(row=r,column=3).value
        is_co=isinstance(av,str) and not str(av).strip().lstrip('-').isdigit() and not cv
        if r==tr: pass
        elif is_co:
            ws.row_dimensions[r].height=20
            for col in range(1,13):
                safe_set(ws.cell(row=r,column=col),fill=NO_FILL,border=b_thin(),
                         font=Font(name='Times New Roman',bold=True,size=12),
                         alignment=Alignment(horizontal='left',vertical='center'))
        else:
            ws.row_dimensions[r].height=bbkn_h(ws,r)
            for col in range(1,13):
                cl=ws.cell(row=r,column=col)
                ha,va=BBKN_A.get(col,('left','center'))
                safe_set(cl,fill=NO_FILL,border=b_thin(),
                         font=Font(name='Times New Roman',size=12),
                         alignment=Alignment(horizontal=ha,vertical=va,wrap_text=col in BBKN_WRAP))
                if col in BBKN_NUM and cl.value is not None: cl.number_format='#,##0'
                if col==8 and isinstance(cl.value,datetime.datetime): cl.number_format='DD/MM/YYYY'

    for col,w in BBKN_W.items(): ws.column_dimensions[get_column_letter(col)].width=w
    ws.page_setup.orientation='landscape'; ws.page_setup.paperSize=ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
    ws.sheet_properties.pageSetUpPr.fitToPage=True
    for a,v in [('left',.4),('right',.4),('top',.5),('bottom',.5),('header',.2),('footer',.2)]:
        setattr(ws.page_margins,a,v)
    ws.print_title_rows='1:14'; ws.freeze_panes=ws.cell(row=DS,column=1)

    out=io.BytesIO(); wb.save(out); out.seek(0); return out.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  MODULE XNT
# ══════════════════════════════════════════════════════════════════════════════
# Template XNT cols: A=STT,B=Tên thuốc,C=Nồng độ,D=ĐVT,E=Số lô,F=Nơi SX,
#                    G=Hạn dùng,H=Đơn giá,I=Tồn đầu,J=Tổng nhập,K=Thực xuất,
#                    L=Tồn cuối,M=Thành tiền,N=Ghi chú
# Raw XNT cols idx: 0=STT,1=?,2=Tên,3=Nồng độ,4=ĐVT,5=Số lô,6=Nơi SX,
#                   7=Hạn dùng,8=Đơn giá,9=Tồn đầu,10=Nhập,11=Xuất,12=Tồn cuối,13=Thành tiền

XNT_W = {1:5,2:28,3:16,4:7,5:10,6:30,7:11,8:12,9:9,10:9,11:9,12:9,13:14,14:8}
XNT_A = {1:('center','center'),2:('left','center'),3:('left','center'),
         4:('center','center'),5:('center','center'),6:('left','center'),
         7:('center','center'),8:('right','center'),9:('right','center'),
         10:('right','center'),11:('right','center'),12:('right','center'),
         13:('right','center'),14:('center','center')}
XNT_WRAP = {2,3,6}
XNT_NUM  = {8,9,10,11,12,13}

def xnt_h(ws,r):
    ml=1
    for c in XNT_WRAP:
        v=ws.cell(row=r,column=c).value
        if not v or not isinstance(v,str): continue
        cw=max(XNT_W.get(c,15)*1.1,1)
        ml=max(ml,sum(max(1,math.ceil(len(ln)/cw)) for ln in v.split('\n')))
    return max(20,min(ml*14.3+4,120))

def build_xnt(tmpl_bytes, companies):
    wb=load_workbook(io.BytesIO(tmpl_bytes))
    ws=wb.active
    cs={c:gs(ws,12,c) for c in range(1,15)}
    ds={c:gs(ws,13,c) for c in range(1,15)}

    # Tìm dòng Tổng cộng gốc trong template -> XÓA đi để tránh duplicate
    fs=None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value=='Tổng cộng': fs=cell.row; break
        if fs: break
    if not fs: fs=279

    ws.delete_rows(fs, 1)   # xóa dòng Tổng cộng gốc của template

    DS=12
    need=sum(1+len(d) for _,d in companies)+1
    data_end=DS+need-1
    ins=data_end-fs+1
    if ins>0: ws.insert_rows(fs,ins); fs+=ins

    for m in [str(mr) for mr in ws.merged_cells.ranges if DS<=mr.min_row<fs]:
        ws.merged_cells.remove(m)
    for r in range(DS,fs):
        for c in range(1,15):
            try: ws.cell(row=r,column=c).value=None
            except: pass

    def wco(rn,name):
        cl=ws.cell(row=rn,column=1,value=name); ap(cl,cs[1])
        cl.font=Font(name='Times New Roman',bold=True,size=11); cl.fill=NO_FILL
        for c in range(2,15):
            cc=ws.cell(row=rn,column=c); ap(cc,cs[c]); cc.fill=NO_FILL
        ws.row_dimensions[rn].height=18

    def wdr(rn,stt,dr):
        cols=[
            (1, stt,                                                    'center',False,None),
            (2, ''if pd.isna(dr[2])else str(dr[2]).strip(),            'left',  True, None),
            (3, ''if pd.isna(dr[3])else str(dr[3]).strip(),            'left',  True, None),
            (4, ''if pd.isna(dr[4])else str(dr[4]).strip(),            'center',False,None),
            (5, ''if pd.isna(dr[5])else str(dr[5]).strip(),            'center',False,None),
            (6, ''if pd.isna(dr[6])else str(dr[6]).strip(),            'left',  True, None),
            (7, dr[7] if isinstance(dr[7],datetime.datetime)
                else(''if pd.isna(dr[7])else dr[7]),                   'center',False,'DD/MM/YYYY'),
            (8, dr[8] if not pd.isna(dr[8])else 0,                    'right', False,'#,##0'),
            (9, dr[9] if not pd.isna(dr[9])else 0,                    'right', False,'#,##0'),
            (10,dr[10]if not pd.isna(dr[10])else 0,                   'right', False,'#,##0'),
            (11,dr[11]if not pd.isna(dr[11])else 0,                   'right', False,'#,##0'),
            (12,dr[12]if not pd.isna(dr[12])else 0,                   'right', False,'#,##0'),
        ]
        for col,val,ha,wrap,fmt in cols:
            cl=ws.cell(row=rn,column=col,value=val); ap(cl,ds[col])
            cl.font=Font(name='Times New Roman',size=11)
            cl.alignment=Alignment(horizontal=ha,vertical='center',wrap_text=wrap)
            if fmt and val!='': cl.number_format=fmt
        cm=ws.cell(row=rn,column=13,value=f'=H{rn}*L{rn}'); ap(cm,ds[13])
        cm.font=Font(name='Times New Roman',size=11)
        cm.alignment=Alignment(horizontal='right',vertical='center')
        cm.number_format='#,##0'
        cn=ws.cell(row=rn,column=14,value=''); ap(cn,ds[14])

    cr=DS; drn=[]
    for name,drugs in companies:
        wco(cr,name); cr+=1
        for i,dr in enumerate(drugs,1): wdr(cr,i,dr); drn.append(cr); cr+=1

    tr=cr
    lbl=ws.cell(row=tr,column=1,value='Tổng cộng')
    lbl.font=Font(name='Times New Roman',bold=True,size=11)
    lbl.alignment=Alignment(horizontal='left',vertical='center'); lbl.border=b_med()
    for c in range(2,13):
        try: ws.cell(row=tr,column=c).border=b_med()
        except: pass
    cm=ws.cell(row=tr,column=13,value=f'=SUM({",".join(f"M{r}"for r in drn)})')
    cm.font=Font(name='Times New Roman',bold=True,size=11)
    cm.alignment=Alignment(horizontal='right',vertical='center')
    cm.number_format='#,##0'; cm.border=b_med()
    ws.row_dimensions[tr].height=20

    # Fix footer: căn giữa chức vụ và tên ký, bỏ wrap text để không xuống dòng
    footer_start_row = tr + 1
    for r in range(footer_start_row, ws.max_row+1):
        for col in range(1,15):
            cl = ws.cell(row=r, column=col)
            if cl.value and isinstance(cl.value, str):
                cl.alignment = Alignment(horizontal='center', vertical='center',
                                         wrap_text=False)
                cl.font = Font(name='Times New Roman', size=11)
        ws.row_dimensions[r].height = 20

    for col in range(1,15):
        for r in range(9,12):
            safe_set(ws.cell(row=r,column=col),fill=NO_FILL,
                     font=Font(name='Times New Roman',bold=True,size=11),
                     border=b_med(),
                     alignment=Alignment(horizontal='center',vertical='center',wrap_text=True))
        ws.row_dimensions[r].height=20

    for r in range(DS,tr+1):
        av=ws.cell(row=r,column=1).value; bv=ws.cell(row=r,column=2).value
        is_co=isinstance(av,str) and not str(av).strip().lstrip('-').isdigit() and not bv
        if r==tr: pass
        elif is_co:
            ws.row_dimensions[r].height=18
            for col in range(1,15):
                safe_set(ws.cell(row=r,column=col),fill=NO_FILL,border=b_thin(),
                         font=Font(name='Times New Roman',bold=True,size=11),
                         alignment=Alignment(horizontal='left',vertical='center'))
        else:
            ws.row_dimensions[r].height=xnt_h(ws,r)
            for col in range(1,15):
                cl=ws.cell(row=r,column=col)
                ha,va=XNT_A.get(col,('left','center'))
                safe_set(cl,fill=NO_FILL,border=b_thin(),
                         font=Font(name='Times New Roman',size=11),
                         alignment=Alignment(horizontal=ha,vertical=va,wrap_text=col in XNT_WRAP))
                if col in XNT_NUM and cl.value is not None and cl.value!='': cl.number_format='#,##0'
                if col==7 and isinstance(cl.value,datetime.datetime): cl.number_format='DD/MM/YYYY'

    for col,w in XNT_W.items(): ws.column_dimensions[get_column_letter(col)].width=w
    ws.page_setup.orientation='landscape'; ws.page_setup.paperSize=ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
    ws.sheet_properties.pageSetUpPr.fitToPage=True
    for a,v in [('left',.35),('right',.35),('top',.5),('bottom',.5),('header',.2),('footer',.2)]:
        setattr(ws.page_margins,a,v)
    ws.print_title_rows='1:11'; ws.freeze_panes=ws.cell(row=DS,column=1)

    out=io.BytesIO(); wb.save(out); out.seek(0); return out.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  GIAO DIỆN
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="hero">
  <div class="badge">🏥 Bệnh viện Đà Nẵng · Khoa Dược</div>
  <h1>HỆ THỐNG TỰ ĐỘNG HÓA<br>BIÊN BẢN DƯỢC</h1>
  <p class="sub">Biên Bản Kiểm Nhập (BBKN) &nbsp;·&nbsp; Báo Cáo Xuất Nhập Tồn (XNT)</p>
</div>
""", unsafe_allow_html=True)

report_type = st.radio(
    "**Chọn loại báo cáo cần xử lý:**",
    ["📋  Biên Bản Kiểm Nhập (BBKN)", "📊  Báo Cáo Xuất Nhập Tồn (XNT)"],
    horizontal=True,
)
is_bbkn = report_type.startswith("📋")

if is_bbkn:
    st.markdown("""<div class="tab-desc">
    Upload <b>file dữ liệu thô BBKN từ HPT</b> (.xls/.xlsx) và <b>file form chuẩn kiểm nhập 2026</b>.<br>
    Logic: Lọc bỏ dòng Số lượng nhập = 0 · Phân nhóm theo công ty · Tính thành tiền tự động.
    </div>""", unsafe_allow_html=True)
    raw_types = ["xls","xlsx"]
    raw_label = "📂 File dữ liệu thô HPT (BBKN)"
    tpl_label = "📄 File form chuẩn Kiểm Nhập 2026"
    qty_col   = 9
    lbls      = ("Công ty cung cấp","Mặt hàng có nhập","Dòng SL=0 đã lọc")
else:
    st.markdown("""<div class="tab-desc">
    Upload <b>file dữ liệu thô XNT từ HPT</b> (.xlsx) và <b>file BBXNT-2026.xlsx</b> làm form chuẩn.<br>
    Logic: Giữ dòng có Tồn cuối ≠ 0 · Phân nhóm theo công ty · Thành tiền = Đơn giá × Tồn cuối.
    </div>""", unsafe_allow_html=True)
    raw_types = ["xlsx"]
    raw_label = "📂 File dữ liệu thô HPT (XNT)"
    tpl_label = "📄 File BBXNT-2026.xlsx (form chuẩn)"
    qty_col   = 12   # cột Tồn cuối
    lbls      = ("Công ty cung cấp","Mặt hàng phát sinh","Dòng tồn cuối=0 đã lọc")

col1,col2 = st.columns(2)
with col1: raw_file  = st.file_uploader(raw_label, type=raw_types)
with col2: tpl_file  = st.file_uploader(tpl_label, type=["xlsx"])

st.markdown("<hr>", unsafe_allow_html=True)

ready = raw_file is not None and tpl_file is not None
if st.button("⚡  Bắt đầu xử lý", disabled=not ready):
    with st.spinner("Đang xử lý dữ liệu..."):
        try:
            raw_b = raw_file.read()
            if raw_file.name.endswith('.xls'):
                try:    raw_df=pd.read_excel(io.BytesIO(raw_b),sheet_name=0,header=None,engine='xlrd')
                except: raw_df=pd.read_excel(io.BytesIO(raw_b),sheet_name=0,header=None)
            else:
                raw_df=pd.read_excel(io.BytesIO(raw_b),sheet_name=0,header=None)

            tpl_b = tpl_file.read()
            companies, stats = parse_companies(raw_df, qty_col)

            if not companies:
                st.error("❌ Không tìm thấy dữ liệu hợp lệ. Kiểm tra lại file HPT.")
                st.stop()

            result = build_bbkn(tpl_b, companies) if is_bbkn else build_xnt(tpl_b, companies)
            st.session_state.update(result=result, stats=stats,
                                    done=True, is_bbkn=is_bbkn, lbls=lbls)
        except Exception as e:
            st.error(f"❌ Lỗi: {e}"); st.exception(e)

if st.session_state.get("done"):
    stats  = st.session_state["stats"]
    result = st.session_state["result"]
    lbls   = st.session_state["lbls"]
    now    = datetime.datetime.now()
    pre    = "BBKN" if st.session_state["is_bbkn"] else "XNT"
    fname  = f"{pre}_T{now.month}_{now.year}_HoanChinh.xlsx"

    st.markdown(f"""
    <div class="ok-box">
      <div class="icon">✅</div>
      <h3>Xử lý hoàn tất!</h3>
      <p>File sẵn sàng — tải về và in ký hội đồng.</p>
    </div>
    <div class="stat-grid">
      <div class="stat-card"><div class="num">{stats['companies']}</div><div class="lbl">{lbls[0]}</div></div>
      <div class="stat-card"><div class="num">{stats['drugs']}</div><div class="lbl">{lbls[1]}</div></div>
      <div class="stat-card"><div class="num">{stats['skipped']}</div><div class="lbl">{lbls[2]}</div></div>
    </div>""", unsafe_allow_html=True)

    st.download_button(
        label=f"⬇️  Tải File Hoàn Chỉnh – {fname}",
        data=result, file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.markdown("""<div class="note">
    💡 <b>Khi in:</b> File đã thiết lập sẵn <b>A4 Ngang · Fit All Columns on One Page</b> · 
    Tiêu đề cột lặp lại mỗi trang. Mở Excel → Ctrl+P → in ngay.
    </div>""", unsafe_allow_html=True)
